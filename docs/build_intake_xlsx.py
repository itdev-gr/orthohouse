#!/usr/bin/env python3
"""
Generate docs/OrthoHouse-Content-Intake.xlsx from the structured content
audit. Run:   python3 docs/build_intake_xlsx.py
Output:      docs/OrthoHouse-Content-Intake.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1E3A5F")
SUBTITLE_FONT = Font(italic=True, size=10, color="555555")
THIN = Side(border_style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

STATUS_OPTIONS = '"Not started,In review,Delivered,N/A"'


def header_row(ws, row, headers, widths=None):
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDER
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.row_dimensions[row].height = 28


def data_rows(ws, start_row, rows):
    for r_offset, row in enumerate(rows):
        for c_offset, val in enumerate(row):
            cell = ws.cell(row=start_row + r_offset, column=c_offset + 1, value=val)
            cell.alignment = WRAP
            cell.border = BORDER


def add_status_validation(ws, col_letter, first_row, last_row):
    dv = DataValidation(type="list", formula1=STATUS_OPTIONS, allow_blank=True)
    dv.error = "Pick: Not started, In review, Delivered, or N/A"
    dv.errorTitle = "Invalid status"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def title_block(ws, title, subtitle=None):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.row_dimensions[1].height = 24
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.row_dimensions[2].height = 18
        return 4  # header row index
    return 3


# ── Workbook ────────────────────────────────────────────────────────────────
wb = Workbook()


# 1) README ─────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "README"
ws["A1"] = "OrthoHouse Cyprus — Content Intake"
ws["A1"].font = Font(bold=True, size=20, color="1E3A5F")
ws["A2"] = "Prepared 19 May 2026 · audit source: docs/CLIENT_CONTENT_INTAKE.md"
ws["A2"].font = SUBTITLE_FONT

readme_lines = [
    "",
    "HOW TO USE THIS WORKBOOK",
    "",
    "This workbook has two purposes:",
    "  1. Document everything the website already pulls live from your Shopify store (sheet: 'Shopify Data').",
    "  2. Track every piece of content / decision we still need from you (sheets A through M plus 'Master List').",
    "",
    "Each action-item sheet has a 'Status' column. Use the dropdown to pick:",
    "  · Not started   — you have not begun gathering this item",
    "  · In review     — being prepared on your side",
    "  · Delivered     — file uploaded / decision sent — we wire it in within 24 h",
    "  · N/A           — does not apply / decided to skip",
    "",
    "The 'Master List' sheet shows every single action item flat in one place — use Excel's filter to",
    "quickly count what is delivered vs outstanding.",
    "",
    "WORKBOOK CONTENTS",
    "",
    "  README           — this page",
    "  Shopify Data     — inventory of fields read live from your Shopify Admin (no action needed)",
    "  Master List      — every action item across all sections, flat (filterable)",
    "  A. Site-wide     — logo, hero copy, hero photo",
    "  B. Shopify Handles — collection / subcategory URL slugs",
    "  C. Banners       — category-page banner photographs + slogans",
    "  D. Supplier Logos — 17 manufacturer logos",
    "  E. Quote-only    — Sensory Rooms + Stair Lifts product catalogue",
    "  F. Product Enrichment — brand, sizing, PDFs, video, FAQs per product",
    "  G. Clarifications — two SKU-specific questions",
    "  H. Home Page Copy — 14 strings shown on the home page",
    "  I. Contact Details — phone / email / address / hours / map",
    "  J. Static Pages  — About / Find Us / Working Hours / Shipping / Refunds / Cookies / Legal / Terms",
    "  K. Business Decisions — newsletter, account, social, GDPR, privacy, chat, analytics, inventory",
    "  L. SEO & Social — meta tags, OG image, favicons, Search Console",
    "  M. Body-Map Subcategories — 20 home-page body-dot handles to verify",
    "",
    "TOTAL: ≈ 150 individual deliverables / confirmations.",
    "",
    "Send files to the shared folder we provided. For each batch, ping us and we'll wire it in within 24 h.",
]
for i, line in enumerate(readme_lines, start=3):
    ws.cell(row=i, column=1, value=line)
ws.column_dimensions["A"].width = 105


# 2) Shopify Data (the "List 1" we already use) ─────────────────────────────
ws = wb.create_sheet("Shopify Data")
hr = title_block(
    ws,
    "List 1 — Fields the website reads live from Shopify (no action needed)",
    "Source of truth: your Shopify Admin. Edit a product/collection in Shopify and the site refreshes.",
)
header_row(ws, hr, ["Group", "Shopify field", "Where used on the website", "Notes"], widths=[24, 50, 50, 50])

shopify_rows = [
    # Per-product
    ["Per product", "product.id", "Internal — Shopify GID for cart calls", ""],
    ["Per product", "product.title", "Product page H1, every card title, breadcrumb leaf, cart line", "UPPERCASE comes from Shopify — you control casing"],
    ["Per product", "product.handle", "URL slug (e.g. /products/prs630), Recently-Viewed entries, static path generation", "If you change a handle in Shopify the live URL changes"],
    ["Per product", "product.description", "Plain-text fallback for card hover & short-desc clamp", ""],
    ["Per product", "product.descriptionHtml", "Product detail 'Description' tab — full rich HTML body", "We strip injected <script> tags as a safety measure"],
    ["Per product", "product.tags", "Future search facets — not yet rendered", ""],
    ["Per product", "product.priceRange.minVariantPrice.amount", "Card price label", ""],
    ["Per product", "product.priceRange.minVariantPrice.currencyCode", "Currency prefix (EUR)", ""],
    ["Per product", "product.featuredImage.url + altText", "Card thumbnail, search overlay, recently-viewed", ""],
    ["Per product", "product.images.nodes[].url + altText", "Product detail gallery + lightbox (first 10 images)", ""],
    ["Per product", "product.availableForSale", "Product-card stock pill (In stock / Out of stock)", "Added per Remarks §9"],
    ["Per product", "product.collections.nodes[].handle + title", "Breadcrumb resolution + quote-only detection (Sensory / Stair Lifts)", "We fetch the first 10 collections per product"],
    ["Per product", "product.variants.nodes[].id", "Variant GID for Add-to-Cart line input", ""],
    ["Per product", "product.variants.nodes[].title", "Variant selector dropdown label (e.g. 'Size M')", ""],
    ["Per product", "product.variants.nodes[].sku", "Product code line on detail page", ""],
    ["Per product", "product.variants.nodes[].availableForSale", "Disables variant option + Add-to-Cart button", ""],
    ["Per product", "product.variants.nodes[].price.amount + currencyCode", "Per-variant price after selector change", ""],
    ["Per product", "product.variants.nodes[].compareAtPrice.amount", "Strike-through 'was' price when on sale", ""],
    ["Per product", "product.variants.nodes[].selectedOptions[]", "Future search facets", ""],
    # Per-collection
    ["Per collection", "collection.id", "Internal cache key", ""],
    ["Per collection", "collection.title", "Collection page H1 (when not overridden by nav menu config)", ""],
    ["Per collection", "collection.handle", "URL slug for /collections/<handle>", ""],
    ["Per collection", "collection.description", "Subcategory page body; junk strings filtered with regex", ""],
    ["Per collection", "collection.image.url + altText", "AVAILABLE but currently unused (we use static banner images)", "Could switch to Shopify-driven banners — see Section K (optional)"],
    ["Per collection", "collection.products.nodes[]", "Card grid on /collections/<handle> (paginated, all pages combined)", ""],
    # Cart
    ["Cart (live)", "cart.id", "LocalStorage-persisted cart token", ""],
    ["Cart (live)", "cart.checkoutUrl", "Checkout button destination → Shopify-hosted checkout", ""],
    ["Cart (live)", "cart.lines.edges[].node.id", "Quantity / remove operations on a cart line", ""],
    ["Cart (live)", "cart.lines.edges[].node.quantity", "Quantity display on /cart", ""],
    ["Cart (live)", "cart.lines.edges[].node.merchandise.id", "Variant GID", ""],
    ["Cart (live)", "cart.lines.edges[].node.merchandise.title", "Variant title on cart line", ""],
    ["Cart (live)", "cart.lines.edges[].node.merchandise.image.url + altText", "Cart-line thumbnail", ""],
    ["Cart (live)", "cart.lines.edges[].node.merchandise.price.amount + currencyCode", "Cart-line price", ""],
    ["Cart (live)", "cart.lines.edges[].node.merchandise.product.title", "Cart-line product name", ""],
    ["Cart (live)", "cart.cost.subtotalAmount.amount + currencyCode", "Cart subtotal footer", ""],
    # Build-time
    ["Build-time only", "All product handles", "Static path generation (/products/<handle>) — ~54 products", "Cached on disk for 5 days"],
    ["Build-time only", "All collection handles", "Static path generation (/collections/<handle>)", ""],
    ["Build-time only", "9-product preview per body-dot subcategory", "Home-page interactive body diagram (20 handles — see sheet M)", ""],
    # Identity
    ["Connection identity", "SHOPIFY_STORE", "orthohouse-connecta.myshopify.com", ""],
    ["Connection identity", "SHOPIFY_API_VERSION", "2025-10", ""],
    ["Connection identity", "STOREFRONT_TOKEN", "Public token in .env", ""],
    ["Connection identity", "SHOPIFY_PRIVATE_TOKEN", "Server-only token in .env", ""],
    # NOT pulled (could be)
    ["NOT used (potential)", "product.metafields.*", "Could drive sizing guides / brand / certs / PDFs from Shopify Admin instead of code", "See Section K8 / optional refactor"],
    ["NOT used (potential)", "collection.metafields.*", "Could let you edit banner image / slogan / supplier list from Shopify", "Optional"],
    ["NOT used (potential)", "Customer accounts / login / order history", "Not implemented", "See Section K2"],
    ["NOT used (potential)", "Reviews / ratings", "Not implemented", ""],
    ["NOT used (potential)", "Inventory quantity (totalInventory / quantityAvailable)", "Only boolean availableForSale is fetched", "See Section K8"],
]
data_rows(ws, hr + 1, shopify_rows)


# Common helper to build a section sheet ────────────────────────────────────
def section_sheet(title_short, title_long, subtitle, rows, columns, col_widths, status_col_letter):
    """
    columns: list of header strings — must include 'Status'
    rows:    list of lists — each row, with values for every column EXCEPT Status (which we leave blank for the user)
    """
    ws = wb.create_sheet(title_short)
    hr = title_block(ws, title_long, subtitle)
    header_row(ws, hr, columns, widths=col_widths)
    for r_offset, row in enumerate(rows):
        # pad with empty string in the Status column position if not present
        for c_offset, val in enumerate(row):
            cell = ws.cell(row=hr + 1 + r_offset, column=c_offset + 1, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
    add_status_validation(ws, status_col_letter, hr + 1, hr + len(rows))
    return ws


# A — Site-wide assets
section_sheet(
    "A. Site-wide",
    "Section A — Site-wide assets",
    "Three top-level deliverables that apply to the whole site.",
    [
        ["A1", "High-resolution logo", "Transparent PNG ≥ 960×240 px, plus an SVG if available",
         "Currently using a low-res file downloaded from the old site as stop-gap", ""],
        ["A2", "'Why customers trust us' revised copy", "Two short paragraphs 60-120 words each (EN + EL), optional new heading",
         "White card on home page below hero", ""],
        ["A3", "Hero photograph", "Single source JPG/PNG ≥ 2400 px wide",
         "Current image works — tell us if you want a different one", ""],
    ],
    columns=["ID", "Item", "Required format", "Notes", "Status"],
    col_widths=[6, 36, 50, 50, 16],
    status_col_letter="E",
)


# B — Shopify collection handles
b_top_rows = [
    ["B-T1", "Top-level", "Health & Medical Devices / Ιατρικές Συσκευές Υγείας", "health-medical-devices", "Missing — create or remap", ""],
    ["B-T2", "Top-level", "Kybun Shoes & Mats / Παπούτσια & Τάπητες Kybun", "kybun-shoes", "Exists in Shopify", ""],
    ["B-T3", "Top-level", "Orthopaedics / Ορθοπεδικά", "orthopedic-braces-supports", "Exists in Shopify", ""],
    ["B-T4", "Top-level", "Home Care & Daily Living / Φροντίδα στο Σπίτι", "home-care-daily-living", "Missing — create or remap", ""],
    ["B-T5", "Top-level", "Health & Comfort / Υγεία & Άνεση", "health-comfort", "Missing — or reuse exercise-and-well-being", ""],
    ["B-T6", "Top-level", "Sensory Rooms / Αισθητηριακά Δωμάτια", "sensory-rooms", "Missing — quote-only collection", ""],
    ["B-T7", "Top-level", "Lifting Solutions / Λύσεις Ανύψωσης", "lifting-solutions", "Missing", ""],
    ["B-T8", "Top-level", "Walking Aids / Βοηθήματα Βάδισης", "walking-aids", "Exists in Shopify", ""],
    ["B-T9", "Top-level", "Wheelchairs / Αναπηρικά Αμαξίδια", "wheelchairs", "Missing — or reuse wheelchairs-and-home-care-aids", ""],
]
b_sub_rows = [
    ["B-S1",  "Subcategory", "Winix - Air Purifiers / Καθαριστές Αέρα", "purifiers", "Verify", ""],
    ["B-S2",  "Subcategory", "Oxygen / Οξυγόνο", "oxygen", "Missing", ""],
    ["B-S3",  "Subcategory", "LifeVac / Συσκευή Αντιμετώπισης Πνιγμού", "lifevac", "Missing", ""],
    ["B-S4",  "Subcategory", "Kybun Men's Shoes / Ανδρικά Παπούτσια", "kybun-mens-shoes", "Missing", ""],
    ["B-S5",  "Subcategory", "Kybun Women's Shoes / Γυναικεία Παπούτσια", "kybun-womens-shoes", "Missing", ""],
    ["B-S6",  "Subcategory", "Kybun Mats / Τάπητες", "kybun-mats", "Missing", ""],
    ["B-S7",  "Subcategory", "Back & Lumbar / Πλάτη & Οσφύς", "back-lumbar", "Consolidate with lumbar / trunk-lumbar-supports", ""],
    ["B-S8",  "Subcategory", "Hospital Beds / Νοσοκομειακά Κρεβάτια", "hospital-beds", "Verify", ""],
    ["B-S9",  "Subcategory", "Bathroom / Μπάνιο", "bathroom-aids", "Verify", ""],
    ["B-S10", "Subcategory", "Toilet Aids / Βοηθήματα Τουαλέτας", "toilets-aids", "Verify", ""],
    ["B-S11", "Subcategory", "Decubitus Prevention / Πρόληψη Κατακλίσεων", "decubitus-aids", "Verify", ""],
    ["B-S12", "Subcategory", "Pillows & Cushions / Μαξιλάρια", "pillows-and-cushions", "Verify", ""],
    ["B-S13", "Subcategory", "Exercise Equipment / Εξοπλισμός Άσκησης", "exercise-equipment", "Missing", ""],
    ["B-S14", "Subcategory", "Massage Equipment / Εξοπλισμός Μασάζ", "massage-equipment", "Missing", ""],
    ["B-S15", "Subcategory", "Hot & Cold Therapy / Θεραπεία Ζεστού & Κρύου", "hot-cold-therapy", "Verify", ""],
    ["B-S16", "Subcategory", "Sensory Rooms Products / Προϊόντα", "sensory-rooms-products", "Missing (quote-only)", ""],
    ["B-S17", "Subcategory", "Sensory Rooms Solutions / Ολοκληρωμένες Λύσεις", "sensory-rooms-solutions", "Missing (quote-only)", ""],
    ["B-S18", "Subcategory", "Orthostats / Ορθοστάτες", "orthostats", "Missing", ""],
    ["B-S19", "Subcategory", "Patient Lifting Hoists / Γερανοί Ανύψωσης", "patient-lifters-hoists", "Verify", ""],
    ["B-S20", "Subcategory", "Stair Lifts by Lehner / Ανελκυστήρες Σκάλας", "stair-lifts-lehner", "Missing (quote-only)", ""],
    ["B-S21", "Subcategory", "Lifting Platforms by Lehner / Πλατφόρμες Ανύψωσης", "lifting-platforms-lehner", "Missing", ""],
    ["B-S22", "Subcategory", "Manual Wheelchairs / Χειροκίνητα", "manual-wheelchairs", "Verify", ""],
    ["B-S23", "Subcategory", "Electric Wheelchairs / Ηλεκτρικά", "wheelchairs-1", "Verify", ""],
]
section_sheet(
    "B. Shopify Handles",
    "Section B — Shopify collection handles (URL slugs)",
    "Top-level + subcategory handles to confirm or create in Shopify Admin. The 9 Orthopaedic subcats and 3 Walking-Aid subcats are already correct.",
    b_top_rows + b_sub_rows,
    columns=["ID", "Tier", "Menu label (EN / EL)", "Expected handle", "Current Shopify status", "Status"],
    col_widths=[8, 12, 50, 35, 38, 16],
    status_col_letter="F",
)


# C — Category banners
c_rows = [
    ["C1", "Health & Medical Devices", "Trusted devices for home", "Αξιόπιστες συσκευές για το σπίτι", "/images/categories/health-medical-devices.webp", "#0f766e", ""],
    ["C2", "Kybun Shoes & Mats", "Walk on air, all day", "Περπατήστε στον αέρα, όλη μέρα", "/images/categories/kybun-shoes.webp", "#334155", ""],
    ["C3", "Orthopaedics", "Support, relief, recovery", "Στήριξη, ανακούφιση, αποκατάσταση", "/images/categories/orthopedics.webp", "#1e3a5f", ""],
    ["C4", "Home Care & Daily Living", "Independence at home", "Ανεξαρτησία στο σπίτι", "/images/categories/home-care.webp", "#7c3aed", ""],
    ["C5", "Health & Comfort", "Rest, recover, feel better", "Ξεκούραση, αποκατάσταση, ευεξία", "/images/categories/health-comfort.webp", "#0369a1", ""],
    ["C6", "Sensory Rooms", "Calming spaces, tailored therapy", "Χώροι ηρεμίας, εξατομικευμένη θεραπεία", "/images/categories/sensory-rooms.webp", "#be185d", ""],
    ["C7", "Lifting Solutions", "Safe transfers, accessible homes", "Ασφαλείς μεταφορές, προσβάσιμα σπίτια", "/images/categories/lifting-solutions.webp", "#b45309", ""],
    ["C8", "Walking Aids", "Confident steps, every day", "Σίγουρα βήματα, κάθε μέρα", "/images/categories/walking-aids.webp", "#047857", ""],
    ["C9", "Wheelchairs", "Mobility without compromise", "Κινητικότητα χωρίς συμβιβασμούς", "/images/categories/wheelchairs.webp", "#374151", ""],
]
section_sheet(
    "C. Banners",
    "Section C — Category page banner content",
    "For each category supply: final EN slogan, final EL slogan, banner photograph (≥ 1600×500 JPG/PNG — we convert to WebP). Reply 'approve current' to keep what's shown.",
    c_rows,
    columns=["ID", "Category", "Current EN slogan", "Current EL slogan", "Current banner image", "Accent colour", "Status"],
    col_widths=[6, 28, 36, 36, 42, 14, 16],
    status_col_letter="G",
)


# D — Supplier logos
d_rows = [
    ["D1",  "Winix", "Health & Medical Devices", "", "Placeholder text-only SVG in code", ""],
    ["D2",  "LifeVac", "Health & Medical Devices", "", "Placeholder", ""],
    ["D3",  "Kybun", "Kybun Shoes & Mats", "", "Confirm placeholder is correct", ""],
    ["D4",  "Bauerfeind", "Orthopaedics", "", "Placeholder", ""],
    ["D5",  "DonJoy", "Orthopaedics", "", "Placeholder", ""],
    ["D6",  "Sponaplast", "Orthopaedics + Walking Aids", "https://www.sponaplast.com", "Confirm — placeholder may be correct", ""],
    ["D7",  "Invacare", "Home Care + Wheelchairs", "", "Placeholder", ""],
    ["D8",  "Etac", "Home Care", "", "Placeholder", ""],
    ["D9",  "Medisana", "Health & Comfort", "", "Placeholder", ""],
    ["D10", "Beurer", "Health & Comfort", "", "Placeholder", ""],
    ["D11", "Rompa", "Sensory Rooms", "", "Placeholder", ""],
    ["D12", "Lehner", "Lifting Solutions", "", "Confirm placeholder is correct", ""],
    ["D13", "Molift", "Lifting Solutions", "", "Placeholder", ""],
    ["D14", "FDI", "Walking Aids", "", "Confirm placeholder is correct", ""],
    ["D15", "AliMed", "Walking Aids", "https://www.alimed.com", "Confirm placeholder is correct", ""],
    ["D16", "Permobil", "Wheelchairs", "", "Placeholder", ""],
    ["D17", "Ottobock", "Wheelchairs", "", "Placeholder", ""],
]
section_sheet(
    "D. Supplier Logos",
    "Section D — Manufacturer / supplier logos",
    "Format: SVG preferred (vector) or transparent PNG ≥ 240 px wide. Adding the brand homepage URL makes the logo clickable. If a supplier is no longer a partner, mark Status = N/A.",
    d_rows,
    columns=["ID", "Supplier", "Used on category", "Homepage URL (optional)", "Notes", "Status"],
    col_widths=[6, 18, 32, 36, 38, 16],
    status_col_letter="F",
)


# E — Quote-only catalogue
e_rows = [
    ["E1", "Sensory Rooms — Products", "sensory-rooms-products", "Each product: title + description + high-res photographs (no price needed — site hides it)", ""],
    ["E2", "Sensory Rooms — Complete Solutions", "sensory-rooms-solutions", "Each entry: title + description + high-res photographs", ""],
    ["E3", "Stair Lifts (Lehner)", "stair-lifts-lehner", "Each product: title + description + high-res photographs", ""],
]
section_sheet(
    "E. Quote-only",
    "Section E — Quote-only catalogue: Sensory Rooms & Stair Lifts",
    "Create real products in Shopify Admin under these collection handles. The site shows 'Contact us for free consultation and quotation' instead of a price.",
    e_rows,
    columns=["ID", "Collection", "Shopify handle", "Per-product requirements", "Status"],
    col_widths=[6, 38, 30, 60, 16],
    status_col_letter="E",
)


# F — Per-product enrichment
f_rows = [
    # Field schema
    ["F-Schema", "Brand", "{ name, logo SVG/PNG ≥ 240 px }", "Renders as a badge on the product page", ""],
    ["F-Schema", "Sizing guide", "HTML snippet OR PDF with measurement chart", "Opens in a modal", ""],
    ["F-Schema", "Financial Assistance PDF", "PDF — GeSY / social-insurance eligibility info", "Eligible-to-Financial-Assistance badge", ""],
    ["F-Schema", "Wolt delivery PDF", "PDF — describing the Wolt delivery process", "A single shared PDF works", ""],
    ["F-Schema", "Certificate logos", "Array of { label, SVG, optional PDF }", "Each clickable — opens its PDF", ""],
    ["F-Schema", "Specifications", "Plain text list of label/value pairs", "Renders as a spec table", ""],
    ["F-Schema", "Feature icons", "SVG ≥ 64 px + 1-2 word label", "Small icon pills (Breathable, Latex-free, etc.)", ""],
    ["F-Schema", "Documents", "PDFs (brochure, user manual, care, declaration of conformity)", "Download-tile grid", ""],
    ["F-Schema", "YouTube video", "YouTube URL — we convert to embed form", "Embedded as a tab", ""],
    ["F-Schema", "Additional content", "HTML or plain text blocks with optional side image", "Standalone content blocks below tabs", ""],
    ["F-Schema", "FAQs", "Question/Answer pairs", "Accordion section", ""],
    # Demo SKUs needing real files
    ["F-kfm2003a", "Airolo Moon Rock M shoes", "8 real PDFs to replace placeholder.pdf", "financial-assistance, ce-cert, swiss-made-cert, wolt-delivery, brochure, size-guide, care, ce-doc", ""],
    ["F-kfm2003a", "Airolo Moon Rock M shoes", "Real YouTube embed URL", "Currently a generic placeholder", ""],
    ["F-sp14029", "Knee support", "6 real PDFs to replace placeholder.pdf", "financial-assistance, ce-cert, wolt-delivery, brochure, ce-cert (second slot), user manual", ""],
    ["F-sp14029", "Knee support", "Real YouTube embed URL", "Currently a Rickroll — MUST REPLACE", ""],
    ["F-the158", "Electric wheelchair", "4 real PDFs to replace placeholder.pdf", "ce-cert, catalogue, ce-cert (second slot), user manual", ""],
    ["F-the158", "Electric wheelchair", "Real YouTube embed URL", "Currently a Rickroll — MUST REPLACE", ""],
    ["F-the158", "Electric wheelchair", "Real Mobiak brand logo SVG", "Brand currently points to OrthoHouse generic logo", ""],
]
section_sheet(
    "F. Product Enrichment",
    "Section F — Per-product enrichment",
    "First rows describe the schema. Below those are concrete files needed for the 3 demo products that are already wired in.",
    f_rows,
    columns=["ID", "Subject", "Item / format", "Notes", "Status"],
    col_widths=[14, 28, 56, 50, 16],
    status_col_letter="E",
)


# G — Clarifications
g_rows = [
    ["G1", "Is product sp14029 (knee support) own-brand OrthoHouse or another manufacturer? Brand record currently set to 'Orthohouse'.", ""],
    ["G2", "Official Mobiak logo SVG for product the158 (electric wheelchair) — currently shows the generic OrthoHouse logo file.", ""],
]
section_sheet(
    "G. Clarifications",
    "Section G — Specific product clarifications",
    "Two open questions on the 3 demo products.",
    g_rows,
    columns=["ID", "Question", "Status"],
    col_widths=[6, 110, 16],
    status_col_letter="C",
)


# H — Home page copy
h_rows = [
    ["H1", "Hero tagline (large text over hero banner)",
     "AT ORTHOHOUSE, WE MAKE IT EASIER FOR YOU TO MOVE BETTER AND LIVE BETTER",
     "ΣΤΗΝ ORTHOHOUSE ΚΑΝΟΥΜΕ ΕΥΚΟΛΟΤΕΡΗ ΤΗΝ ΚΙΝΗΣΗ ΚΑΙ ΤΗ ΖΩΗ ΣΑΣ", ""],
    ["H2", "Intro carousel slide 1 — Orthotics body (~70 words)",
     "Our goal is to help people move more freely and with less pain, correct malpositions and provide effective support for the healing process. Orthoses and supports have proven to be particularly effective for these purposes. From the ground-breaking C-Brace to our complete suite of stroke solutions, we're here to support you.",
     "(Greek equivalent already in code)", ""],
    ["H3", "Intro carousel slide 2 — Wheelchairs body (~60 words)",
     "From manual chairs to powered solutions, we offer a complete range of wheelchairs designed for comfort, stability, and reliability. Whether you need a lightweight transport chair or a fully customized seating system, our team helps you find the right fit for your daily life.",
     "(Greek equivalent already in code)", ""],
    ["H4", "Intro carousel slide 3 — Walking Aids body (~55 words)",
     "Crutches, walkers, rollators, and canes — walking aids give you the confidence to move safely with every step. We curate products that combine ergonomic design, stability, and lightweight construction so you can reclaim your independence both at home and out and about.",
     "(Greek equivalent already in code)", ""],
    ["H5a", "Orthopaedic Showcase title", "Orthopaedic Products", "Ορθοπεδικά Προϊόντα", ""],
    ["H5b", "Orthopaedic Showcase subtitle", "Premium supports & braces for every need", "Κορυφαία στηρίγματα και νάρθηκες για κάθε ανάγκη", ""],
    ["H6a", "GeSY banner — label", "GeSY Approved", "Εγκεκριμένα από ΓεΣΥ", ""],
    ["H6b", "GeSY banner — headline", "Products with GeSY codes available", "Προϊόντα διαθέσιμα με κωδικούς ΓεΣΥ", ""],
    ["H6c", "GeSY banner — sub", "Ask us about health-system covered orthopaedic products", "Ρωτήστε μας για ορθοπεδικά προϊόντα που καλύπτονται από το σύστημα υγείας", ""],
    ["H6d", "GeSY banner — CTA button", "View Full Collection", "Δείτε Όλη τη Συλλογή", ""],
    ["H7",  "'Shop by Category' heading", "Shop by Category", "Αγοράστε ανά Κατηγορία", ""],
    ["H8",  "Category-tile CTA label", "View Solutions", "Δείτε Λύσεις", ""],
    ["H9a", "'New Products' section heading", "New Products", "Νέα Προϊόντα", ""],
    ["H9b", "New-products tabs", "All / Wheelchairs & Home Care / Comfort & Well Being", "Όλα / Αναπηρικά & Φροντίδα / Άνεση & Ευεξία", ""],
]
section_sheet(
    "H. Home Page Copy",
    "Section H — Home page copy",
    "Reply 'approve' per row, or send the replacement EN + EL text.",
    h_rows,
    columns=["ID", "Where", "Current English", "Current Greek", "Status"],
    col_widths=[7, 38, 55, 55, 16],
    status_col_letter="E",
)


# I — Contact details
i_rows = [
    ["I1", "Main phone", "+357 22 66 06 88", "Confirm correct", ""],
    ["I2", "Email", "info@orthohouse.com.cy", "Confirm correct", ""],
    ["I3", "Address EN", "27B 25th March Street, Engomi, 2408 Nicosia, Cyprus", "Confirm correct", ""],
    ["I4", "Address EL", "27B Οδός 25ης Μαρτίου, Έγκωμη, 2408 Λευκωσία, Κύπρος", "Confirm correct", ""],
    ["I5", "Company legal name", "O.H. OrthoHouse Ltd", "Confirm correct", ""],
    ["I6", "Emergency / 24h phone (Working Hours page only)", "+357 99 77 43 09", "Confirm correct, or tell us to remove this section", ""],
    ["I7", "Map coordinates (Google Maps embed + Waze link)", "35.1648933, 33.3309487", "Confirm exact pin location", ""],
    ["I8", "Opening hours", "Mon-Fri 09:00-19:00 · Sat 09:00-14:00 · Sun closed", "Confirm — also if these change on public holidays", ""],
    ["I9 BUG", "Find Us page shows wrong phone (+357 25 123 456)", "We will fix this once you confirm I1", "Data drift to repair", ""],
]
section_sheet(
    "I. Contact Details",
    "Section I — Contact details to confirm",
    "These values appear across many pages. Confirm each and we update once.",
    i_rows,
    columns=["ID", "Item", "Current value", "Action / notes", "Status"],
    col_widths=[10, 36, 55, 50, 16],
    status_col_letter="E",
)


# J — Static page bodies
j_rows = [
    ["J1", "About Us", "/pages/about-us", 647, "First draft — confirm founding year, 'family-run', Engomi-Hippocrateon mention", ""],
    ["J2", "Find Us", "/pages/find-us", 1078, "Likely accurate once I1-I8 confirmed", ""],
    ["J3", "Working Hours", "/pages/working-hours", 739, "Confirm I6 (emergency line) and I8 (hours)", ""],
    ["J4", "Shipping & Delivery", "/pages/shipping-delivery", 1421, "Review against actual shipping practices", ""],
    ["J5", "Refund & Cancellation Policy", "/pages/refund-cancellation-policy", 1030, "Lawyer review recommended", ""],
    ["J6", "Cookies Policy", "/pages/cookies-policy", 1137, "Lawyer review recommended", ""],
    ["J7", "Legal Notice", "/pages/legal-notice", 466, "Confirm legal entity, VAT number, registered office", ""],
    ["J8", "Terms & Conditions", "/pages/terms-and-conditions", 3746, "Lawyer review STRONGLY recommended", ""],
    ["J9", "Terms of Service", "/pages/terms-of-service", 3736, "Lawyer review STRONGLY recommended", ""],
    ["J-VAT", "VAT registration number", "(not currently displayed anywhere)", 0, "Cypriot tax law requires this on legal notice / footer — send the VAT number", ""],
]
section_sheet(
    "J. Static Pages",
    "Section J — Static page bodies (legal / policy review)",
    "Each page was populated with first-draft content. Please review each for legal/factual accuracy and send revised text where needed.",
    j_rows,
    columns=["ID", "Page", "URL", "Word count", "Notes / required action", "Status"],
    col_widths=[8, 38, 38, 12, 60, 16],
    status_col_letter="F",
)


# K — Business decisions
k_rows = [
    ["K1", "Newsletter signup at footer",
     "Connect to Mailchimp / Klaviyo / Shopify Email — OR remove the form. Currently the input does nothing.",
     "URGENT — customer expectations gap", ""],
    ["K2", "Customer accounts",
     "Enable Shopify Customer Accounts (login/orders/addresses) — OR hide the Account person-icon. Currently /account 404s.",
     "URGENT — visible bug", ""],
    ["K3", "Social media",
     "Send Facebook / Instagram / TikTok / LinkedIn / YouTube URLs — OR confirm no public social presence.",
     "Footer has no social icons", ""],
    ["K4", "Cookie-consent banner",
     "Add a GDPR cookie banner (free open-source lib) — OR rely on Cookies Policy page alone (legally weaker).",
     "Recommended", ""],
    ["K5", "Privacy Policy",
     "Send a separate Privacy Policy text (we'll create /pages/privacy-policy) — OR confirm Cookies Policy doubles as your privacy notice.",
     "Translation key already reserved", ""],
    ["K6", "Live chat",
     "Connect Tawk.to / Crisp / Intercom (send embed snippet) — OR skip.",
     "Optional", ""],
    ["K7", "Analytics",
     "Google Analytics 4 (send measurement ID) / Plausible / Fathom — OR skip. Currently no analytics installed.",
     "Recommended for launch", ""],
    ["K8", "Inventory display",
     "Show actual stock quantity ('Only 3 left') when low — OR keep current binary in-stock / out-of-stock pill only.",
     "Optional", ""],
]
section_sheet(
    "K. Decisions",
    "Section K — Business decisions",
    "Not content gaps but product decisions blocking launch / clarity.",
    k_rows,
    columns=["ID", "Decision", "Options", "Priority / notes", "Status"],
    col_widths=[6, 36, 60, 30, 16],
    status_col_letter="E",
)


# L — SEO & social
l_rows = [
    ["L1", "Brand description — one sentence (EN + EL)", "130-160 characters each",
     "Becomes <meta name=description> (Google search snippet)", ""],
    ["L2", "Brand keywords (5-10 phrases each in EN + EL)", "Comma-separated list",
     "Optional meta keywords + signals for AI search", ""],
    ["L3", "Open Graph image", "1200×630 PNG/JPG, ≤ 1 MB",
     "Preview when someone shares a link in WhatsApp / Facebook / Slack", ""],
    ["L4", "Twitter card image (optional, can reuse L3)", "1200×600 PNG/JPG",
     "Twitter / X sharing", ""],
    ["L5", "Favicon variants", "Apple touch icon 180×180 PNG, plus 512×512 PNG for manifest",
     "iPhone home-screen, Android, Windows tile", ""],
    ["L6", "Google Search Console verification", "We send you a TXT DNS record or a .html file to drop in public/",
     "Monitors indexing & errors", ""],
    ["L7", "Sitemap submission — confirm canonical domain", "Confirm https://orthohouse.com.cy",
     "We auto-generate the sitemap", ""],
    ["L8", "Per-page SEO descriptions (optional)", "One sentence each EN + EL",
     "Unique description per page (About / Find Us / each category)", ""],
]
section_sheet(
    "L. SEO & Social",
    "Section L — SEO & social-sharing assets",
    "The site currently has only a <meta name=viewport> tag. Supply the below for proper search rankings + polished sharing.",
    l_rows,
    columns=["ID", "Item", "Format", "What it does", "Status"],
    col_widths=[6, 42, 38, 50, 16],
    status_col_letter="E",
)


# M — Body-map subcategories
m_rows = [
    ["M1",  "Head", "head", "Exists in Shopify", ""],
    ["M2",  "Neck", "cervical", "Exists in Shopify", ""],
    ["M3",  "Shoulder", "shoulder", "Verify", ""],
    ["M4",  "Thoracic (upper back)", "thoracic", "Verify", ""],
    ["M5",  "Elbow", "elbow", "Exists in Shopify", ""],
    ["M6",  "Upper limb", "upper-limb", "Verify", ""],
    ["M7",  "Wrist & Thumb", "wrist-thumb-supports", "Exists in Shopify", ""],
    ["M8",  "Hand", "hand", "Exists in Shopify", ""],
    ["M9",  "Fingers", "fingers", "Verify", ""],
    ["M10", "Lumbar", "lumbar", "Verify — overlaps with back-lumbar and trunk-lumbar-supports", ""],
    ["M11", "Trunk / Lumbar Supports", "trunk-lumbar-supports", "Verify", ""],
    ["M12", "Hip", "hip", "Exists in Shopify", ""],
    ["M13", "Thigh", "thigh", "Verify", ""],
    ["M14", "Lower limb", "lower-limb", "Verify", ""],
    ["M15", "Knee", "knee", "Exists in Shopify", ""],
    ["M16", "Knee braces", "knee-braces", "Verify", ""],
    ["M17", "Calf", "calf", "Verify", ""],
    ["M18", "Ankle", "ankle", "Exists in Shopify", ""],
    ["M19", "Foot", "foot", "Verify", ""],
    ["M20", "Orthopaedic insoles", "orthopedic-insoles", "Verify", ""],
]
section_sheet(
    "M. Body-Map Subcats",
    "Section M — Body-map subcategory handles (home page)",
    "The home page has an interactive human-body diagram (20 clickable dots). Each loads products from a Shopify collection with the handle below. Confirm or remap each — missing handles show an empty grid.",
    m_rows,
    columns=["ID", "Body region", "Expected handle", "Current status", "Status"],
    col_widths=[6, 30, 30, 60, 16],
    status_col_letter="E",
)


# Master List ───────────────────────────────────────────────────────────────
master = wb.create_sheet("Master List", index=2)  # right after Shopify Data
hr = title_block(
    master,
    "Master List — every action item across all sections (filterable)",
    "Use the column filters to quickly count Delivered vs Not started. Each row links to the detailed sheet for its section.",
)
header_row(master, hr, ["ID", "Section", "Item summary", "Required format / action", "Priority", "Status"], widths=[10, 26, 60, 60, 14, 16])

PRIO = {
    "URGENT": "🟥 Urgent",
    "HIGH": "🟧 High",
    "NORM": "🟨 Normal",
    "LOW":  "🟩 Low",
}

master_rows = [
    # A
    ["A1", "A. Site-wide assets", "High-resolution logo", "PNG ≥ 960×240 + SVG if available", PRIO["URGENT"], ""],
    ["A2", "A. Site-wide assets", "Trust info-box revised copy (EN + EL)", "Two 60-120 word paragraphs + heading", PRIO["HIGH"], ""],
    ["A3", "A. Site-wide assets", "Hero photograph approval", "JPG/PNG ≥ 2400 px wide", PRIO["LOW"], ""],
    # B top-level
    *[[r[0], "B. Shopify Handles", r[2], f"Confirm or create handle: {r[3]}", PRIO["URGENT"] if "Missing" in r[4] else PRIO["NORM"], ""] for r in b_top_rows],
    # B subcats
    *[[r[0], "B. Shopify Handles", r[2], f"Confirm or create handle: {r[3]}", PRIO["HIGH"] if "Missing" in r[4] else PRIO["NORM"], ""] for r in b_sub_rows],
    # C
    *[[r[0], "C. Banners", r[1], "Approve EN + EL slogan + banner photograph", PRIO["NORM"], ""] for r in c_rows],
    # D
    *[[r[0], "D. Supplier Logos", r[1], "Send SVG / PNG + optional URL", PRIO["HIGH"], ""] for r in d_rows],
    # E
    ["E1", "E. Quote-only", "Sensory Rooms — Products", "Create products in Shopify (no price)", PRIO["HIGH"], ""],
    ["E2", "E. Quote-only", "Sensory Rooms — Complete Solutions", "Create products in Shopify", PRIO["HIGH"], ""],
    ["E3", "E. Quote-only", "Stair Lifts (Lehner)", "Create products in Shopify", PRIO["HIGH"], ""],
    # F (demo SKU files only — schema is reference)
    ["F-kfm-PDFs", "F. Product Enrichment", "Replace 8 placeholder PDFs on kfm2003a", "Real Kybun PDFs (brochure, certs, care, etc.)", PRIO["HIGH"], ""],
    ["F-kfm-yt",  "F. Product Enrichment", "Replace placeholder YouTube URL on kfm2003a", "Real Kybun video URL", PRIO["NORM"], ""],
    ["F-sp-PDFs", "F. Product Enrichment", "Replace 6 placeholder PDFs on sp14029",  "Real knee-support PDFs", PRIO["HIGH"], ""],
    ["F-sp-yt",   "F. Product Enrichment", "Replace Rickroll YouTube URL on sp14029", "MUST REPLACE — currently Rickroll", PRIO["URGENT"], ""],
    ["F-158-PDFs","F. Product Enrichment", "Replace 4 placeholder PDFs on the158",  "Real wheelchair PDFs", PRIO["HIGH"], ""],
    ["F-158-yt",  "F. Product Enrichment", "Replace Rickroll YouTube URL on the158", "MUST REPLACE — currently Rickroll", PRIO["URGENT"], ""],
    ["F-158-brand", "F. Product Enrichment", "Real Mobiak brand logo SVG", "Currently uses OrthoHouse generic logo", PRIO["NORM"], ""],
    # G
    ["G1", "G. Clarifications", "Is sp14029 own-brand or another manufacturer?", "Confirm", PRIO["LOW"], ""],
    ["G2", "G. Clarifications", "Official Mobiak logo SVG", "Send file", PRIO["NORM"], ""],
    # H
    *[[r[0], "H. Home Page Copy", r[1], "Approve current or send replacement EN + EL", PRIO["NORM"], ""] for r in h_rows],
    # I
    *[[r[0], "I. Contact Details", r[1], "Confirm value", PRIO["URGENT"] if r[0] == "I9 BUG" else PRIO["HIGH"], ""] for r in i_rows],
    # J
    *[[r[0], "J. Static Pages", r[1], "Lawyer review or revised text", PRIO["URGENT"] if "STRONGLY" in r[4] else PRIO["HIGH"], ""] for r in j_rows],
    # K
    *[[r[0], "K. Decisions", r[1], "Pick option / send credentials", PRIO["URGENT"] if "URGENT" in r[3] else PRIO["HIGH"], ""] for r in k_rows],
    # L
    *[[r[0], "L. SEO & Social", r[1], r[2], PRIO["HIGH"] if r[0] in ("L1", "L3", "L5", "L6") else PRIO["NORM"], ""] for r in l_rows],
    # M
    *[[r[0], "M. Body-Map Subcats", r[1], f"Confirm handle: {r[2]}", PRIO["NORM"], ""] for r in m_rows],
]
for r_offset, row in enumerate(master_rows):
    for c_offset, val in enumerate(row):
        cell = master.cell(row=hr + 1 + r_offset, column=c_offset + 1, value=val)
        cell.alignment = WRAP
        cell.border = BORDER
add_status_validation(master, "F", hr + 1, hr + len(master_rows))
master.auto_filter.ref = f"A{hr}:F{hr + len(master_rows)}"


# Reorder sheets so README is first, Shopify Data second, Master List third, then A-M
order = [
    "README", "Shopify Data", "Master List",
    "A. Site-wide", "B. Shopify Handles", "C. Banners", "D. Supplier Logos",
    "E. Quote-only", "F. Product Enrichment", "G. Clarifications",
    "H. Home Page Copy", "I. Contact Details", "J. Static Pages",
    "K. Decisions", "L. SEO & Social", "M. Body-Map Subcats",
]
wb._sheets = [wb[name] for name in order]


# Save ──────────────────────────────────────────────────────────────────────
out = Path(__file__).parent / "OrthoHouse-Content-Intake.xlsx"
wb.save(out)

# Summary print
total_rows = len(master_rows)
print(f"Wrote {out}")
print(f"Sheets: {len(wb.sheetnames)}")
print(f"Master List rows: {total_rows}")
